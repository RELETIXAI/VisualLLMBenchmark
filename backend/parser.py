"""Parse benchmark dataset from Excel (.xlsx) or CSV.

Has a small mtime-aware cache so repeated reads of a 35k-row sheet stay fast.

Expected columns (case-insensitive, flexible):
- image / picture / photo  -> embedded image OR file path/URL
- food / dish / name / description -> ground-truth food name/description
- calories / kcal
- protein / protein_g
- carbs / carbohydrates / carbs_g
- fat / fat_g
- fiber / fiber_g
- sugar / sugar_g
- sodium / sodium_mg
- nutrition (freeform fallback)

For .xlsx with embedded images, images are extracted from anchors and saved
to data/images/<sha1>.png. Each row gets `image_path` populated.
"""
from __future__ import annotations

import hashlib
import io
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

NUTRIENT_KEYS = ["calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "sugar_g", "sodium_mg"]

HEADER_MAP = {
    "calories": ["calories", "kcal", "energy", "cal"],
    "protein_g": ["protein", "protein_g", "proteins", "prot"],
    "carbs_g": ["carbs", "carbohydrates", "carbs_g", "carb"],
    "fat_g": ["fat", "fat_g", "fats", "lipids"],
    "fiber_g": ["fiber", "fibre", "fiber_g"],
    "sugar_g": ["sugar", "sugars", "sugar_g"],
    "sodium_mg": ["sodium", "sodium_mg", "salt"],
    "food": ["food", "dish", "name", "item", "meal"],
    "description": ["description", "explanation", "details", "notes", "desc"],
    "image": ["image", "picture", "photo", "img"],
    "nutrition": ["nutrition", "nutrients", "nutritional", "breakdown"],
    "ingredients": ["ingredients", "ingredient", "components"],
    "health_score": ["healthscore", "health_score", "grade", "healthgrade"],
}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _match_header(header: str) -> str | None:
    h = _norm(header)
    for canon, aliases in HEADER_MAP.items():
        for a in aliases:
            if _norm(a) == h or _norm(a) in h:
                return canon
    return None


def _extract_number(text: Any) -> float | None:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    if isinstance(text, (int, float)):
        return float(text)
    s = str(text)
    m = re.search(r"-?\d+(?:[.,]\d+)?", s)
    if not m:
        return None
    return float(m.group(0).replace(",", "."))


def _parse_ingredients_json(text: Any) -> list[dict]:
    """Parse the WILLMA-style ingredients JSON list.

    Each item canonicalised to:
      {name, quantity, unit, calories, protein_g, carbs_g, fat_g,
       sodium_mg, sugar_g, fiber_g, brand}
    Robust against single-quoted JSON, NULL, and trailing whitespace.
    """
    import json as _json
    if text is None:
        return []
    s = str(text).strip()
    if not s or s.lower() in ("null", "none", "nan"):
        return []
    # Try strict JSON first; fall back to single-quote replacement
    data = None
    for attempt in (s, s.replace("'", '"')):
        try:
            data = _json.loads(attempt)
            break
        except Exception:
            continue
    if not isinstance(data, list):
        return []
    out: list[dict] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        def _f(*keys):
            for k in keys:
                if k in item and item[k] is not None:
                    try:
                        return float(item[k])
                    except (TypeError, ValueError):
                        return None
            return None
        out.append({
            "name":      str(item.get("name") or "").strip(),
            "quantity":  _f("quantity", "qty", "amount"),
            "unit":      str(item.get("unit") or "g").strip(),
            "calories":  _f("calories", "kcal"),
            "protein_g": _f("protein", "proteins"),
            "carbs_g":   _f("carbohydrates", "carbs"),
            "fat_g":     _f("fat", "fats"),
            "sodium_mg": _f("sodium"),
            "sugar_g":   _f("sugar", "sugars"),
            "fiber_g":   _f("fiber", "fibre"),
            "brand":     item.get("brandName") or item.get("brand"),
        })
    return [x for x in out if x["name"]]


def _parse_freeform_nutrition(text: str) -> dict[str, float]:
    """Extract nutrient values from a freeform 'nutrition' string.

    Handles formats like 'calories: 250 kcal, protein 12g, carbs 30g, fat 8g'.
    """
    out: dict[str, float] = {}
    if not text:
        return out
    s = str(text).lower()
    patterns = {
        "calories":  r"(?:calorie|kcal|energy|cal)[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "protein_g": r"protein[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "carbs_g":   r"carb(?:ohydrate)?s?[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "fat_g":     r"fat[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "fiber_g":   r"fib(?:er|re)[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "sugar_g":   r"sugar[^\d-]*(-?\d+(?:[.,]\d+)?)",
        "sodium_mg": r"sodium[^\d-]*(-?\d+(?:[.,]\d+)?)",
    }
    for k, pat in patterns.items():
        m = re.search(pat, s)
        if m:
            out[k] = float(m.group(1).replace(",", "."))
    return out


def _save_image_bytes(data: bytes, images_dir: Path) -> str:
    images_dir.mkdir(parents=True, exist_ok=True)
    h = hashlib.sha1(data).hexdigest()[:16]
    # detect ext from magic bytes
    ext = ".png"
    if data[:3] == b"\xff\xd8\xff":
        ext = ".jpg"
    elif data[:4] == b"\x89PNG":
        ext = ".png"
    elif data[:6] in (b"GIF87a", b"GIF89a"):
        ext = ".gif"
    elif data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        ext = ".webp"
    path = images_dir / f"{h}{ext}"
    if not path.exists():
        path.write_bytes(data)
    return str(path)


def _extract_xlsx_images(ws, images_dir: Path) -> dict[int, str]:
    """Map row index (0-based, data rows) -> image path.

    openpyxl exposes ws._images with anchor info. The anchor row is 0-based
    in xlsx, where row 0 = first row in sheet (i.e. header).
    """
    out: dict[int, str] = {}
    for img in getattr(ws, "_images", []) or []:
        try:
            anchor = img.anchor
            row_idx = None
            if hasattr(anchor, "_from") and anchor._from is not None:
                row_idx = anchor._from.row  # 0-based, 0 = header row
            if row_idx is None:
                continue
            data_row = row_idx - 1  # 0 = first data row (skip header)
            if data_row < 0:
                continue
            blob = img._data() if callable(getattr(img, "_data", None)) else img.ref
            if hasattr(blob, "read"):
                blob = blob.read()
            if isinstance(blob, (bytes, bytearray)):
                out[data_row] = _save_image_bytes(bytes(blob), images_dir)
        except Exception:
            continue
    return out


_PARSE_CACHE: dict = {}


def parse_dataset(file_path: str | Path, images_dir: str | Path = "data/images",
                  image_url_template: str | None = None) -> dict:
    """Returns {'rows': [...], 'columns_detected': {...}, 'n': N}.

    image_url_template: if set, when an image cell holds an opaque id
    (not a URL, not an existing file path), build the URL via
    template.format(image_id=<value>) or simple {id} placeholder substitution.
    Example: "https://bucket.example.com/imgs/{image_id}"
    """
    file_path = Path(file_path)
    images_dir = Path(images_dir)
    suffix = file_path.suffix.lower()

    cache_key = (str(file_path), str(image_url_template))
    try:
        mtime = file_path.stat().st_mtime
    except FileNotFoundError:
        mtime = 0
    cached = _PARSE_CACHE.get(cache_key)
    if cached and cached["mtime"] == mtime:
        return cached["result"]

    image_by_row: dict[int, str] = {}
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        image_by_row = _extract_xlsx_images(ws, images_dir)
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            return {"rows": [], "columns_detected": {}, "n": 0}
        headers = [str(c) if c is not None else "" for c in rows_data[0]]
        body = rows_data[1:]
        df = pd.DataFrame(body, columns=headers)
    elif suffix == ".csv":
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    col_map: dict[str, str] = {}
    for col in df.columns:
        canon = _match_header(col)
        if canon and canon not in col_map:
            col_map[canon] = col

    rows = []
    for i, raw in df.iterrows():
        row = {"row_idx": int(i), "image_path": None, "image_url": None,
               "image_id": None, "food": None, "description": None,
               "nutrition_truth": {}, "ingredients_truth": [],
               "health_score_truth": None, "raw": {}}

        if "image" in col_map:
            v = raw.get(col_map["image"])
            if isinstance(v, str) and v.strip():
                v = v.strip()
                if v.startswith(("http://", "https://")):
                    row["image_url"] = v
                elif Path(v).exists():
                    row["image_path"] = v
                elif image_url_template:
                    row["image_url"] = (image_url_template
                                        .replace("{image_id}", v)
                                        .replace("{id}", v))
                    row["image_id"] = v
                else:
                    row["image_id"] = v  # opaque id, no resolver
        if i in image_by_row and not row["image_path"] and not row["image_url"]:
            row["image_path"] = image_by_row[i]

        if "food" in col_map:
            v = raw.get(col_map["food"])
            if v is not None and str(v).strip() and str(v).lower() != "nan":
                row["food"] = str(v).strip()
        if "description" in col_map:
            v = raw.get(col_map["description"])
            if v is not None and str(v).strip() and str(v).lower() != "nan":
                row["description"] = str(v).strip()

        nut: dict[str, float] = {}
        for k in NUTRIENT_KEYS:
            if k in col_map:
                n = _extract_number(raw.get(col_map[k]))
                if n is not None:
                    nut[k] = n
        if "nutrition" in col_map:
            txt = raw.get(col_map["nutrition"])
            if txt is not None:
                for k, v in _parse_freeform_nutrition(str(txt)).items():
                    nut.setdefault(k, v)
        row["nutrition_truth"] = nut

        if "ingredients" in col_map:
            row["ingredients_truth"] = _parse_ingredients_json(raw.get(col_map["ingredients"]))
        if "health_score" in col_map:
            v = raw.get(col_map["health_score"])
            if v is not None and str(v).strip() and str(v).strip().lower() != "nan":
                row["health_score_truth"] = str(v).strip().upper()[:1] if str(v).strip()[0:1].isalpha() else str(v).strip()

        row["raw"] = {k: (str(v) if v is not None and str(v).lower() != "nan" else None)
                      for k, v in raw.items()}
        rows.append(row)

    result = {"rows": rows, "columns_detected": col_map, "n": len(rows)}
    _PARSE_CACHE[cache_key] = {"mtime": mtime, "result": result}
    return result
